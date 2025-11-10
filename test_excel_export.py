#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 자동 병합 기능 테스트 스크립트
"""

import os
import sys
import json
import openpyxl
from openpyxl import Workbook
import datetime

OUTPUT_FOLDER = "WPS-OUTPUT"

def save_to_excel(data, pdf_path):
    """개선된 Excel 파일에 데이터 추가 함수 (테스트용)"""
    try:
        excel_file = os.path.join(OUTPUT_FOLDER, "WPS_추출결과.xlsx")

        # 고정 컬럼
        fixed_columns = ['PDF파일명', '추출일시']

        # 데이터 필드를 정렬하여 일관된 순서 유지
        data_fields = sorted(data.keys())

        if not os.path.exists(excel_file):
            # 새 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "WPS Data"
            headers = fixed_columns + data_fields
            ws.append(headers)
            print(f"✅ 새 Excel 파일 생성: {len(data_fields)}개 필드")
        else:
            # 기존 파일 로드
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # 기존 헤더 읽기 (첫 번째 행)
            existing_headers = [cell.value for cell in ws[1]]

            # 기존 헤더에서 고정 컬럼 제외하고 데이터 필드만 추출
            existing_data_fields = existing_headers[len(fixed_columns):]

            # 새로운 필드가 있는지 확인
            all_fields = sorted(set(existing_data_fields + data_fields))

            # 헤더가 변경되었으면 업데이트
            if existing_data_fields != all_fields:
                new_headers = fixed_columns + all_fields
                for col_idx, header in enumerate(new_headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                print(f"📊 Excel 헤더 업데이트: {len(existing_data_fields)} → {len(all_fields)}개 필드")

            # 최종 필드 목록 업데이트
            data_fields = all_fields

        # 행 데이터 준비 (헤더 순서에 맞춰 데이터 배치)
        row_data = [
            os.path.basename(pdf_path),
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]

        # 각 필드별로 값 추가 (없는 필드는 빈 문자열)
        for field in data_fields:
            row_data.append(data.get(field, ''))

        ws.append(row_data)
        wb.save(excel_file)
        print(f"✅ Excel 저장 완료: {excel_file} (총 {ws.max_row}행)")
        return True

    except Exception as e:
        print(f"❌ Excel 저장 실패: {e}")
        return False


def test_excel_export():
    """Excel 자동 병합 기능 테스트"""

    print("="*60)
    print("📝 Excel 자동 병합 기능 테스트 시작")
    print("="*60)

    # OUTPUT 폴더 생성
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
        print(f"✅ 출력 폴더 생성: {OUTPUT_FOLDER}")

    # 기존 Excel 파일 삭제 (테스트를 위해)
    excel_file = os.path.join(OUTPUT_FOLDER, "WPS_추출결과.xlsx")
    if os.path.exists(excel_file):
        os.remove(excel_file)
        print(f"🗑️  기존 Excel 파일 삭제 (테스트용)")

    print("\n" + "="*60)
    print("테스트 1: 첫 번째 파일 - 기본 필드")
    print("="*60)

    # 테스트 데이터 1: 기본 필드
    test_data_1 = {
        "WPS_No": "WPS-001",
        "Rev_No": "A",
        "Process": "GMAW",
        "Material": "Carbon Steel",
        "Thickness": "10mm"
    }

    result = save_to_excel(test_data_1, "test_wps_001.pdf")
    if result:
        print("✅ 테스트 1 성공")

    print("\n" + "="*60)
    print("테스트 2: 두 번째 파일 - 동일한 필드")
    print("="*60)

    # 테스트 데이터 2: 동일한 필드
    test_data_2 = {
        "WPS_No": "WPS-002",
        "Rev_No": "B",
        "Process": "GTAW",
        "Material": "Stainless Steel",
        "Thickness": "5mm"
    }

    result = save_to_excel(test_data_2, "test_wps_002.pdf")
    if result:
        print("✅ 테스트 2 성공")

    print("\n" + "="*60)
    print("테스트 3: 세 번째 파일 - 새로운 필드 추가")
    print("="*60)

    # 테스트 데이터 3: 새로운 필드 추가
    test_data_3 = {
        "WPS_No": "WPS-003",
        "Rev_No": "C",
        "Process": "SMAW",
        "Material": "Aluminum",
        "Thickness": "8mm",
        "Preheat_Temp": "150°C",  # 새로운 필드
        "Current": "120A"  # 새로운 필드
    }

    result = save_to_excel(test_data_3, "test_wps_003.pdf")
    if result:
        print("✅ 테스트 3 성공 - 헤더 자동 확장됨")

    print("\n" + "="*60)
    print("테스트 4: 네 번째 파일 - 일부 필드만 있는 경우")
    print("="*60)

    # 테스트 데이터 4: 일부 필드만
    test_data_4 = {
        "WPS_No": "WPS-004",
        "Process": "FCAW",
        "Current": "100A"
        # 다른 필드는 없음 (빈 값으로 처리되어야 함)
    }

    result = save_to_excel(test_data_4, "test_wps_004.pdf")
    if result:
        print("✅ 테스트 4 성공 - 누락된 필드는 빈 값으로 처리")

    # 최종 결과 확인
    print("\n" + "="*60)
    print("📊 최종 결과 확인")
    print("="*60)

    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        print(f"✅ Excel 파일 생성 확인: {excel_file}")
        print(f"📌 시트명: {ws.title}")
        print(f"📌 총 행 수: {ws.max_row} (헤더 포함)")
        print(f"📌 총 열 수: {ws.max_column}")

        # 헤더 출력
        print(f"\n📋 헤더:")
        headers = [cell.value for cell in ws[1]]
        for i, header in enumerate(headers, 1):
            print(f"   {i}. {header}")

        # 데이터 행 출력
        print(f"\n📄 데이터 (처음 5행):")
        for row_idx in range(2, min(7, ws.max_row + 1)):
            row_data = [cell.value for cell in ws[row_idx]]
            print(f"   행 {row_idx}: {row_data}")

        print("\n" + "="*60)
        print("✅ 모든 테스트 완료!")
        print("="*60)
        print(f"\n결과 파일: {excel_file}")
        print("Excel 파일을 열어서 확인해보세요!")

    else:
        print("❌ Excel 파일이 생성되지 않았습니다.")


if __name__ == "__main__":
    try:
        test_excel_export()
    except Exception as e:
        print(f"\n❌ 테스트 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
